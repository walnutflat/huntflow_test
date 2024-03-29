# -*- coding: utf-8 -*-

"""Функции работы с xlsx.
"""

from openpyxl import load_workbook

from db_script import config
from db_script.candidate import Candidate

candidate_fields = ('last_name', 'first_name', 'middle_name')


def parse_db_file(logger, filename: str) -> list:
    """Парсить один файл XLSX.
    """
    # noinspection PyBroadException
    try:
        wb = load_workbook(filename)
    except Exception as exc:
        logger.error(f'Ошибка открытия файла: {exc}')
        return []

    res = []
    sheet = wb['Лист1']
    i = 2
    while True:
        # noinspection PyBroadException
        try:
            candidate_name = sheet.cell(row=i, column=config.DB_NAME_COL).value
            if not candidate_name:
                break

            candidate_name = candidate_name.strip()
            candidate_name_splitted = candidate_name.split(' ')
            candidate_name_dict = {}
            for num, field in enumerate(candidate_fields):
                try:
                    candidate_name_dict[field] = candidate_name_splitted[num]
                except IndexError:
                    candidate_name_dict[field] = None

            new_candidate = Candidate(
                last_name=candidate_name_dict['last_name'],
                first_name=candidate_name_dict['first_name'],
                middle_name=candidate_name_dict['middle_name'],
                position=sheet.cell(row=i, column=config.DB_POSITION_COL).value.strip() if \
                    sheet.cell(row=i, column=config.DB_POSITION_COL).value else None,
                salary=str(sheet.cell(row=i, column=config.DB_SALARY_COL).value).strip() if \
                    sheet.cell(row=i, column=config.DB_SALARY_COL).value else None,
                comment=sheet.cell(row=i, column=config.DB_COMMENT_COL).value.strip() if \
                    sheet.cell(row=i, column=config.DB_COMMENT_COL).value else None,
                status=sheet.cell(row=i, column=config.DB_STATUS_COL).value.strip() if \
                    sheet.cell(row=i, column=config.DB_STATUS_COL).value else None,
                db_name=candidate_name,
            )
            logger.info(f'Обработан кандиат {candidate_name}')
            res.append(new_candidate)
        except Exception as exc:
            logger.error(f'Сбой при обработке кандидата: {exc}')
        i += 1

    wb.close()

    return res

if __name__ == '__main__':
    parse_db_file(None, 'D:/Тестовое задание/Тестовая база.xlsx')
